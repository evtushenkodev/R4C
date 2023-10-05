from datetime import timedelta

from django.db.models import Count
from django.utils import timezone
from openpyxl.workbook import Workbook

from robots.models import Robot


def generate_report():
    # Calculate the date one week ago from now
    one_week_ago = timezone.now() - timedelta(days=7)

    # Create an Excel workbook
    workbook = create_excel_workbook()

    # Get and filter robot data for the past week
    robot_data = filter_robot_data(one_week_ago)

    # Process the robot data and add it to the sheets
    for item in robot_data:
        sheet = get_or_create_sheet(workbook, item['model'])
        if 'A1' not in sheet:
            initialize_sheet(sheet)

        add_data_to_sheet(sheet, item)

    # Remove the default 'Sheet'
    remove_default_sheet(workbook)

    # Return the workbook and the robot data
    return workbook, robot_data


def create_excel_workbook():
    # Create a new Excel workbook
    return Workbook()


def filter_robot_data(one_week_ago):
    # Filter robot data created within the last week
    return (
        Robot.objects.filter(created__gte=one_week_ago)
        .values('model', 'version')
        .annotate(count=Count('id'))
        .distinct()
    )


def get_or_create_sheet(workbook, model):
    # Get or create a sheet with the given model name
    return (
        workbook.get_sheet_by_name(model)
        if model in workbook.sheetnames
        else workbook.create_sheet(model)
    )


def initialize_sheet(sheet):
    # Initialize the headers in the sheet
    sheet['A1'] = 'Model'
    sheet['B1'] = 'Version'
    sheet['C1'] = 'Count for the week'


def add_data_to_sheet(sheet, item):
    # Add data to the sheet
    row = sheet.max_row + 1
    sheet.cell(row=row, column=1, value=item['model'])
    sheet.cell(row=row, column=2, value=item['version'])
    sheet.cell(row=row, column=3, value=item['count'])


def remove_default_sheet(workbook):
    # Remove the default 'Sheet'
    workbook.remove(workbook['Sheet'])